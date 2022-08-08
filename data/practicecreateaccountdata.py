import openpyxl


class CreateAccountData:
    # createAccountData = [{"Your name": "janani", "Mobile number": 123456, "Email": "123@123.com", "Password": 123456}]
    @staticmethod
    def getExcelData(test_case_name):
        dataDict = {}
        wbObj = openpyxl.load_workbook("C:\\Users\\dines\\OneDrive\\Documents\\MyPythonProject\\DataSetUp.xlsx")
        wsObj = wbObj.active
        for i in range(1, wsObj.max_row+1):
            if wsObj.cell(row=i, column=1).value == test_case_name:
                for j in range(2, wsObj.max_column+1):
                    dataDict[wsObj.cell(row=1, column=j).value] = wsObj.cell(row=i, column=j).value

        return [dataDict]


